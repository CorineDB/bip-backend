#!/usr/bin/env python3
"""
Script pour créer un template propre sans les questions pré-remplies
Supprime les lignes 14-49 et enlève toutes les cellules fusionnées dans cette zone
"""

from openpyxl import load_workbook

template_path = '/home/unknow/GDIZ/apps/backend_api/canevas/O-5_Outil_evaluation_de_la_note_conceptuelle.xlsx'
clean_template_path = '/home/unknow/GDIZ/apps/backend_api/canevas/O-5_Template_Appreciation_Clean.xlsx'

print("="*80)
print("CRÉATION DU TEMPLATE PROPRE")
print("="*80)

# Charger le template original
wb = load_workbook(template_path)
sheet = wb.active

print("\n✓ Template original chargé")

# Trouver toutes les cellules fusionnées dans la zone 14-49
merged_to_remove = []
for merged_range in list(sheet.merged_cells.ranges):
    # Vérifier si cette fusion touche les lignes 14-49
    if any(row in range(14, 50) for row in range(merged_range.min_row, merged_range.max_row + 1)):
        merged_to_remove.append(merged_range)

print(f"\n✓ Trouvé {len(merged_to_remove)} cellules fusionnées à supprimer")

# Défusionner toutes ces cellules
for merged_range in merged_to_remove:
    sheet.unmerge_cells(str(merged_range))

print(f"✓ Cellules défusionnées")

# Supprimer les lignes 14 à 49 (36 lignes)
# IMPORTANT: Supprimer de la fin vers le début pour éviter les décalages
print(f"\n✓ Suppression des lignes 14-49...")
sheet.delete_rows(14, 36)

print(f"✓ Lignes supprimées")

# IMPORTANT: Après suppression, les lignes de conclusion sont maintenant à partir de ligne 14
# Il faut AUSSI défusionner toutes les cellules qui sont maintenant >= ligne 14
print(f"\n✓ Défusion des cellules restantes après ligne 13...")
remaining_merged = list(sheet.merged_cells.ranges)
merged_to_remove_2 = []
for merged_range in remaining_merged:
    if merged_range.min_row >= 14:
        merged_to_remove_2.append(merged_range)

print(f"✓ Trouvé {len(merged_to_remove_2)} cellules à défusionner après ligne 13")

for merged_range in merged_to_remove_2:
    sheet.unmerge_cells(str(merged_range))

print(f"✓ Toutes les cellules >= ligne 14 sont maintenant défusionnées")

# Vérifier qu'il ne reste pas de cellules fusionnées problématiques
remaining_merged = list(sheet.merged_cells.ranges)
print(f"\n✓ Cellules fusionnées restantes: {len(remaining_merged)}")
for merged in remaining_merged[:10]:  # Afficher les 10 premières
    print(f"  - {merged}")

# Sauvegarder le template propre
wb.save(clean_template_path)

print(f"\n{'='*80}")
print(f"✓ TEMPLATE PROPRE CRÉÉ: {clean_template_path}")
print(f"{'='*80}")
print(f"\nLigne 14 est maintenant prête à recevoir les sections et questions dynamiques")
print(f"La section CONCLUSION commence maintenant à la ligne 14 (au lieu de 50)")
print("="*80)
