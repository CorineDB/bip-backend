#!/usr/bin/env python3
"""
Script pour générer le fichier Excel d'appréciation de la note conceptuelle
Reproduit EXACTEMENT la structure du template avec tous les styles
Usage: python3 generate_appreciation_excel.py <data_json_file> <template_path> <output_path>
"""

import sys
import json
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

def create_section_row(sheet, row, title):
    """
    Créer une ligne de section (fond vert FF09A493)
    Structure: A:D fusionnées, E avec même fond
    """
    # Fusionner A:D
    sheet.merge_cells(f'A{row}:D{row}')

    # Cellule A (titre de la section)
    cell_a = sheet[f'A{row}']
    cell_a.value = title
    cell_a.font = Font(bold=True, size=16, color='FFFFC000')  # Orange/Or
    cell_a.fill = PatternFill(start_color='FF09A493', end_color='FF09A493', fill_type='solid')  # Vert
    cell_a.alignment = Alignment(vertical='center', wrap_text=True)
    cell_a.border = Border(
        left=Side(style='medium'),
        bottom=Side(style='thin')
    )

    # Cellules B, C, D (fusionnées avec A, mais on applique les bordures)
    for col in ['B', 'C', 'D']:
        cell = sheet[f'{col}{row}']
        cell.border = Border(bottom=Side(style='thin'))

    # Cellule E (même fond vert)
    cell_e = sheet[f'E{row}']
    cell_e.font = Font(bold=True, size=16, color='FFFFC000')
    cell_e.fill = PatternFill(start_color='FF09A493', end_color='FF09A493', fill_type='solid')
    cell_e.alignment = Alignment(vertical='center', wrap_text=True)
    cell_e.border = Border(bottom=Side(style='thin'))

    # Hauteur de ligne
    sheet.row_dimensions[row].height = 27.0

def create_question_rows(sheet, row, question_data):
    """
    Créer 2 lignes pour une question
    Ligne 1: Titre (A:C), Commentaire (D sur 2 lignes), Guide (E sur 2 lignes)
    Ligne 2: Vide (A:B), Validation (C)
    """
    row1 = row
    row2 = row + 1

    # ========== LIGNE 1 ==========

    # Fusionner A:C pour le titre
    sheet.merge_cells(f'A{row1}:C{row1}')
    cell_a = sheet[f'A{row1}']
    cell_a.value = question_data.get('title', '')
    cell_a.font = Font(bold=True, size=11)
    cell_a.alignment = Alignment(vertical='top', wrap_text=True)
    cell_a.border = Border(left=Side(style='medium'), top=Side(style='thin'))

    # Bordures pour B et C (fusionnés avec A)
    sheet[f'B{row1}'].border = Border(top=Side(style='thin'))
    sheet[f'C{row1}'].border = Border(top=Side(style='thin'))

    # Colonne D: Commentaire (fusionné sur 2 lignes, fond gris)
    # IMPORTANT: Définir valeur et styles AVANT de fusionner
    cell_d = sheet[f'D{row1}']
    cell_d.value = question_data.get('comment', '')
    cell_d.font = Font(size=10)
    cell_d.fill = PatternFill(start_color='FFF3F3F3', end_color='FFF3F3F3', fill_type='solid')  # Gris
    cell_d.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    cell_d.border = Border(
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    sheet.merge_cells(f'D{row1}:D{row2}')

    # Colonne E: Guide (fusionné sur 2 lignes)
    # IMPORTANT: Définir valeur et styles AVANT de fusionner
    cell_e = sheet[f'E{row1}']
    cell_e.value = question_data.get('guide', '')
    cell_e.font = Font(size=10)
    cell_e.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    cell_e.border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    sheet.merge_cells(f'E{row1}:E{row2}')

    # Hauteur ligne 1
    sheet.row_dimensions[row1].height = 20.25

    # ========== LIGNE 2 ==========

    # Fusionner A:B (vide)
    sheet.merge_cells(f'A{row2}:B{row2}')
    cell_a2 = sheet[f'A{row2}']
    cell_a2.value = ''
    cell_a2.font = Font(size=11)
    cell_a2.alignment = Alignment(horizontal='center', vertical='top', wrap_text=True)
    cell_a2.border = Border(
        left=Side(style='medium'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Bordure pour B (fusionné avec A)
    sheet[f'B{row2}'].border = Border(top=Side(style='thin'), bottom=Side(style='thin'))

    # Colonne C: Bouton de validation avec appréciation (fond cyan)
    cell_c = sheet[f'C{row2}']
    appreciation = question_data.get('appreciation', '')

    # Mapper les valeurs d'appréciation
    appreciation_map = {
        'passe': 'Validé',
        'retour': 'Réservé',
        'non_accepte': 'Rejeté',
    }

    appreciation_text = appreciation_map.get(appreciation, '[ Valider le statut ]')
    cell_c.value = appreciation_text
    cell_c.font = Font(bold=True, size=12, color='FF000000')
    cell_c.fill = PatternFill(start_color='FFEBFFFC', end_color='FFEBFFFC', fill_type='solid')  # Cyan
    cell_c.alignment = Alignment(vertical='top', wrap_text=True)

    # Bordures pour D et E (déjà fusionnés avec ligne au-dessus)
    sheet[f'D{row2}'].border = Border(right=Side(style='thin'), bottom=Side(style='thin'))
    sheet[f'E{row2}'].border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Hauteur ligne 2
    sheet.row_dimensions[row2].height = 85.5

def create_accept_text_row(sheet, row, accept_text):
    """
    Créer la ligne de termes de référence (accept_text)
    Ligne unique avec fusion A:E, texte wrap, hauteur 58.5
    """
    # Fusionner A:E
    sheet.merge_cells(f'A{row}:E{row}')

    # Cellule A avec le texte
    cell_a = sheet[f'A{row}']
    cell_a.value = accept_text
    cell_a.font = Font(bold=False, size=11)
    cell_a.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
    cell_a.border = Border(
        left=Side(style='medium'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Hauteur de ligne
    sheet.row_dimensions[row].height = 58.5

def create_proposant_section(sheet, row, proposant_data):
    """
    Créer la section 'À remplir par le proposant du projet'
    Ligne 1: Titre (A:E, fond cyan FFEBFFFC, texte vert FF09A493)
    Ligne 2-3: Informations du proposant (Nom, Téléphone, Email, Ministère)
    """
    row1 = row
    row2 = row + 1
    row3 = row + 2

    # ========== LIGNE 1: Titre de la section ==========
    sheet.merge_cells(f'A{row1}:E{row1}')
    cell_a1 = sheet[f'A{row1}']
    cell_a1.value = "À remplir par le proposant du projet"
    cell_a1.font = Font(bold=True, size=14, color='FF09A493')  # Texte vert
    cell_a1.fill = PatternFill(start_color='FFEBFFFC', end_color='FFEBFFFC', fill_type='solid')  # Fond cyan
    cell_a1.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell_a1.border = Border(
        left=Side(style='medium'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    sheet.row_dimensions[row1].height = 25.5

    # ========== LIGNE 2: Nom, Téléphone, Email ==========
    # Colonne A: "Proposition de projet préparée par (Nom) :"
    cell_a2 = sheet[f'A{row2}']
    nom_proposant = proposant_data.get('nom', '')
    cell_a2.value = f"Proposition de projet préparée par (Nom) :\n{nom_proposant}"
    cell_a2.font = Font(bold=True, size=12, color='FF222A35')
    cell_a2.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
    cell_a2.border = Border(left=Side(style='medium'), top=Side(style='thin'))

    # Colonne B: Téléphone (fusionné sur 2 lignes)
    sheet.merge_cells(f'B{row2}:B{row3}')
    cell_b2 = sheet[f'B{row2}']
    telephone = proposant_data.get('telephone', '')
    cell_b2.value = f"Téléphone:\n{telephone}" if telephone else "Téléphone:"
    cell_b2.font = Font(bold=True, size=12, color='FF222A35')
    cell_b2.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
    cell_b2.border = Border(top=Side(style='thin'), bottom=Side(style='thin'))

    # Colonne C:E fusionnées - Email (sur 2 lignes)
    sheet.merge_cells(f'C{row2}:E{row3}')
    cell_c2 = sheet[f'C{row2}']
    email = proposant_data.get('email', '')
    cell_c2.value = f"E-mail:\n{email}" if email else "E-mail:"
    cell_c2.font = Font(bold=True, size=12, color='FF222A35')
    cell_c2.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
    cell_c2.border = Border(
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    sheet.row_dimensions[row2].height = 18.0

    # ========== LIGNE 3: Nom du ministère ==========
    cell_a3 = sheet[f'A{row3}']
    cell_a3.value = f"Nom du ministère :\n{proposant_data.get('ministere', '')}"
    cell_a3.font = Font(bold=True, size=12, color='FF222A35')
    cell_a3.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
    cell_a3.border = Border(
        left=Side(style='medium'),
        bottom=Side(style='thin')
    )

    # Bordures pour B et C:E déjà fusionnées
    sheet[f'B{row3}'].border = Border(bottom=Side(style='thin'))
    sheet[f'C{row3}'].border = Border(right=Side(style='thin'), bottom=Side(style='thin'))

    sheet.row_dimensions[row3].height = 30.0

def create_resultats_section(sheet, row, last_question_row):
    """
    Créer la section 'Résultats de l'examen'
    Ligne 1: Titre (A:E, fond vert, texte blanc)
    Lignes 2-4: Compteurs avec formules COUNTIF
    Ligne 5: "Le résultat de l'examen est donc le suivant :"
    Lignes 6-9: Options de décision avec champs raisons/recommandations
    """
    row1 = row      # Titre
    row2 = row + 1  # Validées
    row3 = row + 2  # Réservées
    row4 = row + 3  # Rejetées
    row5 = row + 4  # "Le résultat..."
    row6 = row + 5  # Vide
    row7 = row + 6  # Note conforme
    row8 = row + 7  # Vide
    row9 = row + 8  # Avis réservé
    row10 = row + 9 # Raisons et recommandations
    row11 = row + 10 # Vide
    row12 = row + 11 # Note non conforme

    # ========== LIGNE 1: Titre ==========
    sheet.merge_cells(f'A{row1}:E{row1}')
    cell_a1 = sheet[f'A{row1}']
    cell_a1.value = "Résultats de l'examen"
    cell_a1.font = Font(bold=True, size=14, color='FFFFFFFF')  # Texte blanc
    cell_a1.fill = PatternFill(start_color='FF09A493', end_color='FF09A493', fill_type='solid')  # Fond vert
    cell_a1.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell_a1.border = Border(
        left=Side(style='medium'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    sheet.row_dimensions[row1].height = 34.5

    # ========== LIGNE 2: Nombre de rubriques validées ==========
    cell_a2 = sheet[f'A{row2}']
    cell_a2.value = "Nombre de rubriques validées"
    cell_a2.font = Font(bold=True, size=13)
    cell_a2.alignment = Alignment(horizontal='left', vertical='center')
    cell_a2.border = Border(left=Side(style='medium'))

    cell_b2 = sheet[f'B{row2}']
    cell_b2.value = f'=COUNTIF(C$14:C${last_question_row},"Validé")'
    cell_b2.font = Font(bold=True, size=13)
    cell_b2.alignment = Alignment(horizontal='left', vertical='center')

    sheet.row_dimensions[row2].height = 24.75

    # ========== LIGNE 3: Nombre de rubriques réservées ==========
    cell_a3 = sheet[f'A{row3}']
    cell_a3.value = "Nombre de rubriques ayant fait objet de réserve"
    cell_a3.font = Font(bold=True, size=13)
    cell_a3.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    cell_a3.border = Border(left=Side(style='medium'))

    cell_b3 = sheet[f'B{row3}']
    cell_b3.value = f'=COUNTIF(C$14:C${last_question_row},"Réservé")'
    cell_b3.font = Font(bold=True, size=13)
    cell_b3.alignment = Alignment(horizontal='left', vertical='center')

    sheet.row_dimensions[row3].height = 15.75

    # ========== LIGNE 4: Nombre de rubriques rejetées ==========
    cell_a4 = sheet[f'A{row4}']
    cell_a4.value = "Nombre de rubriques rejetées"
    cell_a4.font = Font(bold=True, size=13)
    cell_a4.alignment = Alignment(horizontal='left', vertical='center')
    cell_a4.border = Border(left=Side(style='medium'))

    cell_b4 = sheet[f'B{row4}']
    cell_b4.value = f'=COUNTIF(C$14:C${last_question_row},"Rejeté")'
    cell_b4.font = Font(bold=True, size=13)
    cell_b4.alignment = Alignment(horizontal='left', vertical='center')

    sheet.row_dimensions[row4].height = 24.75

    # ========== LIGNE 5: "Le résultat de l'examen..." ==========
    sheet.merge_cells(f'A{row5}:B{row5 + 1}')
    cell_a5 = sheet[f'A{row5}']
    cell_a5.value = "Le résultat de l'examen est donc le suivant :"
    cell_a5.font = Font(bold=True, size=13)
    cell_a5.fill = PatternFill(start_color='FF09A493', end_color='FF09A493', fill_type='solid')
    cell_a5.alignment = Alignment(horizontal='left', vertical='center')
    cell_a5.border = Border(left=Side(style='medium'))

    sheet.row_dimensions[row5].height = 15.0

    # ========== LIGNE 7: Note conceptuelle conforme ==========
    sheet.merge_cells(f'A{row7}:B{row7 + 1}')
    cell_a7 = sheet[f'A{row7}']
    cell_a7.value = "( ) Note conceptuelle conforme (toutes les rubriques validées)"
    cell_a7.font = Font(bold=True, size=12, color='FF00B050')  # Vert
    cell_a7.alignment = Alignment(horizontal='left', vertical='center')
    cell_a7.border = Border(left=Side(style='medium'))

    sheet.row_dimensions[row7].height = 24.75

    # ========== LIGNE 9: Avis réservé ==========
    sheet.merge_cells(f'A{row9}:B{row9}')
    cell_a9 = sheet[f'A{row9}']
    cell_a9.value = "( ) Avis réservé sur la note conceptuelle (avis réservé)"
    cell_a9.font = Font(bold=True, size=12, color='FFED7D31')  # Orange
    cell_a9.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    cell_a9.border = Border(left=Side(style='medium'))

    sheet.row_dimensions[row9].height = 30.0

    # ========== LIGNE 10: Raisons et recommandations ==========
    sheet.merge_cells(f'A{row10}:B{row10 + 1}')
    cell_a10 = sheet[f'A{row10}']
    cell_a10.value = "Raisons et recommandations d'amélioration :"
    cell_a10.font = Font(bold=True, size=12, color='FFED7D31')  # Orange
    cell_a10.alignment = Alignment(horizontal='left', vertical='top')
    cell_a10.border = Border(left=Side(style='medium'))

    sheet.row_dimensions[row10].height = 48.0

    # ========== LIGNE 12: Note non conforme ==========
    cell_a12 = sheet[f'A{row12}']
    cell_a12.value = "( ) Note conceptuelle non conforme (rejet)"
    cell_a12.font = Font(bold=True, size=12, color='FFFF0000')  # Rouge
    cell_a12.alignment = Alignment(horizontal='left', vertical='center')
    cell_a12.border = Border(left=Side(style='medium'))

    sheet.row_dimensions[row12].height = 15.0

def main():
    if len(sys.argv) != 4:
        print("Usage: python3 generate_appreciation_excel.py <data_json> <template_path> <output_path>")
        sys.exit(1)

    data_json_file = sys.argv[1]
    template_path = sys.argv[2]
    output_path = sys.argv[3]

    # Charger les données JSON
    with open(data_json_file, 'r', encoding='utf-8') as f:
        data = json.load(f)

    # Charger le template PROPRE (sans questions pré-remplies)
    wb = load_workbook(template_path)
    sheet = wb.active

    # Remplir les informations d'en-tête
    header = data.get('header', {})
    if 'titre_projet' in header:
        sheet['B4'] = header['titre_projet']
    if 'identifiant_bip' in header:
        sheet['B5'] = header['identifiant_bip']
    if 'cout_total' in header:
        sheet['B6'] = header['cout_total']
    if 'date_demarrage' in header:
        sheet['B7'] = header['date_demarrage']
    if 'date_achevement' in header:
        sheet['B8'] = header['date_achevement']

    # Point d'insertion: ligne 14 (maintenant vide dans le template propre)
    current_row = 14
    elements = data.get('elements', [])

    # Ajouter les sections et questions dynamiquement
    # Pas besoin d'insert_rows car le template est maintenant propre
    for element in elements:
        element_type = element.get('type')

        if element_type == 'section':
            create_section_row(sheet, current_row, element.get('title', ''))
            current_row += 1

        elif element_type == 'question':
            create_question_rows(sheet, current_row, element)
            current_row += 2

    # La dernière ligne de question pour les formules
    last_question_row = current_row - 1

    # Ajouter la ligne de termes de référence (accept_text)
    accept_text = data.get('accept_text', "En remplissant et en transmettant cette note conceptuelle de projet, je confirme que les informations sont complètes et exactes à ma connaissance. Je reconnais également que le fait de fournir intentionnellement des informations inexactes ou trompeuses peut entraîner des sanctions à mon encontre.")
    create_accept_text_row(sheet, current_row, accept_text)
    current_row += 1  # La ligne accept_text prend 1 ligne

    # Ajouter la section "À remplir par le proposant du projet"
    proposant_data = data.get('proposant', {})
    create_proposant_section(sheet, current_row, proposant_data)
    current_row += 3  # La section proposant prend 3 lignes

    # Ajouter la section "Résultats de l'examen"
    create_resultats_section(sheet, current_row, last_question_row)
    current_row += 12  # La section résultats prend 12 lignes

    # Sauvegarder
    wb.save(output_path)
    print(f"✓ Fichier Excel généré: {output_path}")

    # Retourner des statistiques
    stats = {
        'sections': sum(1 for e in elements if e['type'] == 'section'),
        'questions': sum(1 for e in elements if e['type'] == 'question'),
        'total_rows': current_row - 14,
        'last_question_row': last_question_row
    }
    print(json.dumps(stats))

if __name__ == '__main__':
    main()
