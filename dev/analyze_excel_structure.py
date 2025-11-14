#!/usr/bin/env python3
"""Analyse détaillée de la structure du template Excel"""

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border

template_path = '/home/unknow/GDIZ/apps/backend_api/canevas/O-5_Outil_evaluation_de_la_note_conceptuelle.xlsx'

wb = load_workbook(template_path)
sheet = wb.active

print("="*80)
print("ANALYSE DÉTAILLÉE DU TEMPLATE EXCEL")
print("="*80)

# Analyser l'en-tête (lignes 1-13)
print("\n" + "="*80)
print("SECTION 1: EN-TÊTE (lignes 1-13)")
print("="*80)

for row in range(1, 14):
    has_content = False
    for col in ['A', 'B', 'C', 'D', 'E', 'F']:
        cell = sheet[f'{col}{row}']
        if cell.value:
            has_content = True
            break

    if has_content:
        print(f"\n--- LIGNE {row} ---")
        for col in ['A', 'B', 'C', 'D', 'E', 'F']:
            cell = sheet[f'{col}{row}']
            if cell.value:
                val = str(cell.value)[:50] + "..." if len(str(cell.value)) > 50 else str(cell.value)
                print(f"  {col}{row}: {val}")

                # Styles
                if cell.font:
                    print(f"    Font: Bold={cell.font.bold}, Size={cell.font.size}, Color={cell.font.color.rgb if cell.font.color else 'None'}")
                if cell.fill and cell.fill.start_color:
                    print(f"    Fill: {cell.fill.start_color.rgb}")
                if cell.alignment:
                    print(f"    Align: H={cell.alignment.horizontal}, V={cell.alignment.vertical}, Wrap={cell.alignment.wrap_text}")

# Analyser les sections et questions (lignes 23-29)
print("\n" + "="*80)
print("SECTION 2: TEMPLATE DE SECTIONS ET QUESTIONS (lignes 23-29)")
print("="*80)

for row in range(23, 30):
    print(f"\n{'='*80}")
    print(f"LIGNE {row}")
    print('='*80)

    for col in ['A', 'B', 'C', 'D', 'E', 'F']:
        cell = sheet[f'{col}{row}']
        coord = f'{col}{row}'

        # Valeur
        val = str(cell.value)[:60] if cell.value else "(vide)"
        print(f"\n  {coord}: {val}")

        # Font
        if cell.font:
            font_details = []
            if cell.font.bold: font_details.append("Bold")
            if cell.font.italic: font_details.append("Italic")
            if cell.font.size: font_details.append(f"Size={cell.font.size}")
            if cell.font.color and cell.font.color.rgb:
                font_details.append(f"Color={cell.font.color.rgb}")
            if font_details:
                print(f"    Font: {', '.join(font_details)}")

        # Fill
        if cell.fill and cell.fill.start_color and cell.fill.start_color.rgb:
            rgb = cell.fill.start_color.rgb
            if rgb not in ['00000000', 'FFFFFFFF']:  # Ignorer noir et blanc par défaut
                print(f"    Fill: {rgb}")

        # Alignment
        if cell.alignment:
            align_details = []
            if cell.alignment.horizontal:
                align_details.append(f"H={cell.alignment.horizontal}")
            if cell.alignment.vertical:
                align_details.append(f"V={cell.alignment.vertical}")
            if cell.alignment.wrap_text:
                align_details.append("WrapText=True")
            if align_details:
                print(f"    Align: {', '.join(align_details)}")

        # Border
        if cell.border:
            border_details = []
            if cell.border.left and cell.border.left.style:
                border_details.append(f"Left={cell.border.left.style}")
            if cell.border.right and cell.border.right.style:
                border_details.append(f"Right={cell.border.right.style}")
            if cell.border.top and cell.border.top.style:
                border_details.append(f"Top={cell.border.top.style}")
            if cell.border.bottom and cell.border.bottom.style:
                border_details.append(f"Bottom={cell.border.bottom.style}")
            if border_details:
                print(f"    Border: {', '.join(border_details)}")

# Cellules fusionnées pour lignes 23-29
print("\n" + "="*80)
print("CELLULES FUSIONNÉES (lignes 23-29)")
print("="*80)

merged_ranges = list(sheet.merged_cells.ranges)
relevant_merges = [m for m in merged_ranges if any(f'{r}' in str(m) for r in range(23, 30))]

for merge in relevant_merges:
    print(f"  {merge}")

# Dimensions des colonnes
print("\n" + "="*80)
print("DIMENSIONS DES COLONNES")
print("="*80)

for col in ['A', 'B', 'C', 'D', 'E', 'F']:
    dim = sheet.column_dimensions[col]
    print(f"  Colonne {col}: width={dim.width}")

# Hauteur des lignes 23-29
print("\n" + "="*80)
print("HAUTEUR DES LIGNES (23-29)")
print("="*80)

for row in range(23, 30):
    dim = sheet.row_dimensions[row]
    print(f"  Ligne {row}: height={dim.height}")

print("\n" + "="*80)
print("FIN DE L'ANALYSE")
print("="*80)
