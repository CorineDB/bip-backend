#!/usr/bin/env python3
"""
Script pour tester le clonage de tous les types d'éléments:
- Section principale (ligne unique, fond vert)
- Question (2 lignes: titre + validation)
"""

from openpyxl import load_workbook
from copy import copy

template_path = '/home/unknow/GDIZ/apps/backend_api/canevas/O-5_Outil_evaluation_de_la_note_conceptuelle.xlsx'
output_path = '/home/unknow/GDIZ/apps/backend_api/storage/app/test_excel_clone_all.xlsx'

wb = load_workbook(template_path)
sheet = wb.active

print("=== TEST CLONAGE TOUS TYPES ===\n")

def copy_cell_style(source_cell, target_cell):
    """Copier le style d'une cellule"""
    if source_cell.has_style:
        target_cell.font = copy(source_cell.font)
        target_cell.fill = copy(source_cell.fill)
        target_cell.border = copy(source_cell.border)
        target_cell.alignment = copy(source_cell.alignment)
        target_cell.number_format = copy(source_cell.number_format)
        target_cell.protection = copy(source_cell.protection)

def clone_section(sheet, source_row, target_row, title):
    """
    Cloner une section principale (ligne 23)
    Structure: A:D fusionnées (Bold, Fond vert 09A493), E vide
    """
    print(f"Clonage section: ligne {source_row} → {target_row}")

    # Insérer 1 ligne
    sheet.insert_rows(target_row, 1)

    # Copier styles de toutes les cellules
    for col in ['A', 'B', 'C', 'D', 'E', 'F']:
        source_cell = sheet[f'{col}{source_row}']
        target_cell = sheet[f'{col}{target_row}']
        copy_cell_style(source_cell, target_cell)

    # Fusionner A:D et remplir le titre
    sheet.merge_cells(f'A{target_row}:D{target_row}')
    sheet[f'A{target_row}'].value = title

    print(f"  ✓ Section '{title}' créée à la ligne {target_row}\n")
    return target_row + 1

def clone_question(sheet, source_row_1, source_row_2, target_row, title, comment, guide):
    """
    Cloner une question (lignes 24-25)
    Structure:
    - Ligne 1: Titre (A:C), Commentaire (D:D+1), Guide (E:E+1)
    - Ligne 2: Vide (A:B), Validation (C), Commentaire (D), Guide (E)
    """
    print(f"Clonage question: lignes {source_row_1}-{source_row_2} → {target_row}-{target_row+1}")

    # Insérer 2 lignes
    sheet.insert_rows(target_row, 2)

    # LIGNE 1: Copier styles
    for col in ['A', 'B', 'C', 'D', 'E', 'F']:
        source_cell = sheet[f'{col}{source_row_1}']
        target_cell = sheet[f'{col}{target_row}']
        copy_cell_style(source_cell, target_cell)

    # Fusionner A:C pour le titre
    sheet.merge_cells(f'A{target_row}:C{target_row}')
    sheet[f'A{target_row}'].value = title

    # Fusionner D sur 2 lignes pour le commentaire
    sheet.merge_cells(f'D{target_row}:D{target_row+1}')
    sheet[f'D{target_row}'].value = comment

    # Fusionner E sur 2 lignes pour le guide
    sheet.merge_cells(f'E{target_row}:E{target_row+1}')
    sheet[f'E{target_row}'].value = guide

    # LIGNE 2: Copier styles
    for col in ['A', 'B', 'C', 'D', 'E', 'F']:
        source_cell = sheet[f'{col}{source_row_2}']
        target_cell = sheet[f'{col}{target_row+1}']
        copy_cell_style(source_cell, target_cell)

    # Fusionner A:B (vide)
    sheet.merge_cells(f'A{target_row+1}:B{target_row+1}')
    sheet[f'A{target_row+1}'].value = ''

    # Cellule C: [ Valider le statut ]
    sheet[f'C{target_row+1}'].value = '[ Valider le statut ]'

    print(f"  ✓ Question '{title}' créée aux lignes {target_row}-{target_row+1}\n")
    return target_row + 2

# Test: Ajouter une section et 2 questions à partir de la ligne 30
current_row = 30

# 1. Section principale
current_row = clone_section(
    sheet,
    source_row=23,  # Ligne template de section
    target_row=current_row,
    title="SECTION DE TEST - Nouvelle section d'évaluation"
)

# 2. Question 1
current_row = clone_question(
    sheet,
    source_row_1=24,  # Ligne template question titre
    source_row_2=25,  # Ligne template question validation
    target_row=current_row,
    title="Question test 1: Analyse des risques",
    comment="Commentaire pour l'analyse des risques",
    guide="Validé: Les risques sont bien identifiés et atténués\nRéservé: Certains risques nécessitent clarification\nRejeté: L'analyse des risques est insuffisante"
)

# 3. Question 2
current_row = clone_question(
    sheet,
    source_row_1=24,
    source_row_2=25,
    target_row=current_row,
    title="Question test 2: Budget prévisionnel",
    comment="Commentaire sur le budget",
    guide="Validé: Le budget est détaillé et justifié\nRéservé: Le budget nécessite des ajustements\nRejeté: Le budget est incomplet ou irréaliste"
)

# Sauvegarder
wb.save(output_path)
print(f"✓ Fichier sauvegardé: {output_path}")

print("\n=== RÉSUMÉ ===")
print(f"Section ajoutée à la ligne 30")
print(f"Question 1 ajoutée aux lignes 31-32")
print(f"Question 2 ajoutée aux lignes 33-34")
print(f"\nOuvrir le fichier pour vérifier que:")
print("  - La section a le fond vert (09A493)")
print("  - Les questions ont les bonnes fusions de cellules")
print("  - Les boutons de validation ont le fond cyan (EBFFFC)")
print("  - Les commentaires ont le fond gris (F3F3F3)")
