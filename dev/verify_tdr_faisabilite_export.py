#!/usr/bin/env python3
"""
Vérifier le contenu du fichier Excel généré pour TDR Faisabilité
"""
from openpyxl import load_workbook

file_path = '/home/unknow/GDIZ/apps/backend_api/storage/app/private/projets/936e4225ac474c61e24ee9296f235eedeb49696f049b702127c32f8d3ae06fdb/evaluation_ex_ante/etude_faisabilite/tdr/appreciation_tdr_faisabilite_BIP-2025-6SB2QE.xlsx'

try:
    wb = load_workbook(file_path)
    sheet = wb.active

    print(f'Feuille: {sheet.title}')
    print(f'Dimensions: {sheet.max_row} lignes x {sheet.max_column} colonnes')
    print('\n' + '='*80)
    print('Contenu du fichier (50 premières lignes):')
    print('='*80 + '\n')

    for row_idx in range(1, min(51, sheet.max_row + 1)):
        row_content = []
        for col_idx in range(1, min(6, sheet.max_column + 1)):
            cell = sheet.cell(row_idx, col_idx)
            if cell.value:
                row_content.append(str(cell.value)[:60])

        if row_content:
            print(f'Ligne {row_idx:2d}: {" | ".join(row_content)}')

    print('\n' + '='*80)
    total_rows = sum(1 for row in sheet.iter_rows() if any(cell.value for cell in row))
    print(f'Total de lignes avec contenu: {total_rows}')
    print('='*80)

except Exception as e:
    print(f'Erreur: {e}')
    import traceback
    traceback.print_exc()
