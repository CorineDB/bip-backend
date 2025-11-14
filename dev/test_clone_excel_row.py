#!/usr/bin/env python3
"""
Script pour cloner une ligne de question du template Excel
Structure : 2 lignes par question
- Ligne 1 (ex: 24) : Titre (A24:C24), Commentaire vide (D24:D25), Guide (E24:E25)
- Ligne 2 (ex: 25) : Vide (A25:B25), Validation (C25), Commentaire (D25 fusionné), Guide (E25 fusionné)
"""

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from copy import copy

# Charger le template
template_path = '/home/unknow/GDIZ/apps/backend_api/canevas/O-5_Outil_evaluation_de_la_note_conceptuelle.xlsx'
output_path = '/home/unknow/GDIZ/apps/backend_api/storage/app/test_excel_clone.xlsx'

wb = load_workbook(template_path)
sheet = wb.active

print("=== CLONAGE D'UNE QUESTION ===\n")

# On va cloner les lignes 24-25 (Démarche administrative) pour créer une nouvelle question
# On va l'insérer juste après, aux lignes 30-31

SOURCE_ROW_1 = 24  # Ligne du titre
SOURCE_ROW_2 = 25  # Ligne de validation
TARGET_ROW = 30    # Où insérer la nouvelle question

# Données de test
NEW_TITLE = "Ma nouvelle question test"
NEW_COMMENT = "Ceci est un commentaire de test"
NEW_GUIDE = "Validé: Acceptable\nRéservé: À revoir\nRejeté: Non acceptable"

print(f"Source: lignes {SOURCE_ROW_1}-{SOURCE_ROW_2}")
print(f"Cible: lignes {TARGET_ROW}-{TARGET_ROW+1}")
print(f"Titre: {NEW_TITLE}")
print(f"Commentaire: {NEW_COMMENT}")
print()

# Insérer 2 nouvelles lignes
sheet.insert_rows(TARGET_ROW, 2)
print(f"✓ Inséré 2 lignes à la position {TARGET_ROW}")

# Fonction pour copier le style d'une cellule
def copy_cell_style(source_cell, target_cell):
    if source_cell.has_style:
        target_cell.font = copy(source_cell.font)
        target_cell.fill = copy(source_cell.fill)
        target_cell.border = copy(source_cell.border)
        target_cell.alignment = copy(source_cell.alignment)
        target_cell.number_format = copy(source_cell.number_format)
        target_cell.protection = copy(source_cell.protection)

# LIGNE 1 (Titre + Guide)
print(f"\n--- Copie de la ligne {SOURCE_ROW_1} vers {TARGET_ROW} ---")

# Copier les styles de toutes les cellules de la ligne 1
for col in ['A', 'B', 'C', 'D', 'E', 'F']:
    source_cell = sheet[f'{col}{SOURCE_ROW_1}']
    target_cell = sheet[f'{col}{TARGET_ROW}']
    copy_cell_style(source_cell, target_cell)
    print(f"  ✓ Style {col}{SOURCE_ROW_1} → {col}{TARGET_ROW}")

# Fusionner A:C pour le titre (comme A24:C24)
sheet.merge_cells(f'A{TARGET_ROW}:C{TARGET_ROW}')
sheet[f'A{TARGET_ROW}'].value = NEW_TITLE
print(f"  ✓ Fusionné A{TARGET_ROW}:C{TARGET_ROW} avec titre: '{NEW_TITLE}'")

# Fusionner D sur 2 lignes pour le commentaire (comme D24:D25)
sheet.merge_cells(f'D{TARGET_ROW}:D{TARGET_ROW+1}')
sheet[f'D{TARGET_ROW}'].value = NEW_COMMENT
print(f"  ✓ Fusionné D{TARGET_ROW}:D{TARGET_ROW+1} avec commentaire")

# Fusionner E sur 2 lignes pour le guide (comme E24:E25)
sheet.merge_cells(f'E{TARGET_ROW}:E{TARGET_ROW+1}')
sheet[f'E{TARGET_ROW}'].value = NEW_GUIDE
print(f"  ✓ Fusionné E{TARGET_ROW}:E{TARGET_ROW+1} avec guide")

# LIGNE 2 (Validation)
print(f"\n--- Copie de la ligne {SOURCE_ROW_2} vers {TARGET_ROW+1} ---")

# Copier les styles de toutes les cellules de la ligne 2
for col in ['A', 'B', 'C', 'D', 'E', 'F']:
    source_cell = sheet[f'{col}{SOURCE_ROW_2}']
    target_cell = sheet[f'{col}{TARGET_ROW+1}']
    copy_cell_style(source_cell, target_cell)
    print(f"  ✓ Style {col}{SOURCE_ROW_2} → {col}{TARGET_ROW+1}")

# Fusionner A:B (vide) pour la ligne de validation
sheet.merge_cells(f'A{TARGET_ROW+1}:B{TARGET_ROW+1}')
sheet[f'A{TARGET_ROW+1}'].value = ''
print(f"  ✓ Fusionné A{TARGET_ROW+1}:B{TARGET_ROW+1} (vide)")

# Cellule C : [ Valider le statut ]
sheet[f'C{TARGET_ROW+1}'].value = '[ Valider le statut ]'
print(f"  ✓ C{TARGET_ROW+1} = '[ Valider le statut ]'")

# D et E sont déjà fusionnés avec la ligne au-dessus

# Sauvegarder
wb.save(output_path)
print(f"\n✓ Fichier sauvegardé: {output_path}")

print("\n=== VÉRIFICATION ===")
print(f"Ouvrir le fichier et vérifier les lignes {TARGET_ROW}-{TARGET_ROW+1}")
print(f"- Titre: '{NEW_TITLE}' doit être en A{TARGET_ROW}:C{TARGET_ROW}")
print(f"- Commentaire: '{NEW_COMMENT}' doit être en D{TARGET_ROW}:D{TARGET_ROW+1}")
print(f"- Guide: '{NEW_GUIDE}' doit être en E{TARGET_ROW}:E{TARGET_ROW+1}")
print(f"- Validation: '[ Valider le statut ]' doit être en C{TARGET_ROW+1}")
