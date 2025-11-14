#!/usr/bin/env python3
"""
Script pour vérifier le fichier Excel généré avec resultat_global = 'passe'
"""
import sys
from openpyxl import load_workbook

file_path = "/home/unknow/GDIZ/apps/backend_api/storage/app/private/projets/90ec386312360c44510d5f8e4a7175837a20731745a253c35adfb3e1d642232d/evaluation_ex_ante/etude_profil/note_conceptuelle/appreciation_note_conceptuelle_BIP-2025-D8YMWS.xlsx"

wb = load_workbook(file_path)
sheet = wb.active

print(f"Feuille: {sheet.title}")
print(f"\nRecherche de la section 'Résultats de l'examen'...\n")

in_results_section = False
results_start_row = None

for row_idx in range(1, sheet.max_row + 1):
    cell_value = sheet[f'A{row_idx}'].value

    if cell_value and isinstance(cell_value, str):
        if "Résultats de l'examen" in cell_value:
            in_results_section = True
            results_start_row = row_idx
            print(f"{'='*70}")
            print(f"DÉBUT DE LA SECTION RÉSULTATS (ligne {row_idx})")
            print(f"{'='*70}")

        if in_results_section and results_start_row:
            # Afficher les 20 premières lignes après le début de la section
            if row_idx <= results_start_row + 20:
                # Vérifier si c'est une des options de décision
                if "(X)" in cell_value or "( )" in cell_value:
                    if "(X)" in cell_value:
                        print(f"Ligne {row_idx}: ✓ COCHÉE - {cell_value}")
                    else:
                        print(f"Ligne {row_idx}: ☐ Non cochée - {cell_value}")
                else:
                    print(f"Ligne {row_idx}: {cell_value}")

            # S'arrêter après 25 lignes pour éviter d'afficher trop
            if row_idx > results_start_row + 25:
                break

print(f"\n{'='*70}")
print("Vérification:")

# Vérifier qu'il n'y a que l'option "conforme" cochée (passe)
has_conforme_checked = False
has_reserve_option = False
has_non_conforme_option = False

for row_idx in range(results_start_row, min(results_start_row + 25, sheet.max_row + 1)):
    cell_value = sheet[f'A{row_idx}'].value
    if cell_value and isinstance(cell_value, str):
        if "(X) Note conceptuelle conforme" in cell_value:
            has_conforme_checked = True
        if "Avis réservé" in cell_value and "recommandation" not in cell_value.lower():
            has_reserve_option = True
        if "Note conceptuelle non conforme" in cell_value:
            has_non_conforme_option = True

print(f"✓ Option 'Conforme' cochée: {has_conforme_checked}")
print(f"✗ Ligne 'Avis réservé' présente: {has_reserve_option}")
print(f"✗ Ligne 'Non conforme' présente: {has_non_conforme_option}")

if has_conforme_checked and not has_reserve_option and not has_non_conforme_option:
    print(f"\n✅ SUCCÈS - Seule l'option 'Conforme' est affichée (resultat_global = 'passe')")
else:
    print(f"\n❌ ÉCHEC - Le fichier contient des lignes non attendues")

print(f"{'='*70}\n")
